from fastapi import FastAPI, UploadFile, File, HTTPException, Query
from fastapi.responses import FileResponse, HTMLResponse, Response
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fastapi import Request
from pathlib import Path
from typing import List
from io import BytesIO
import tempfile
import os

from app.models import (
    Quota, QuotaUpdate, ProjectUsage, 
    ReportFile, ImportResult,
    K8sConfig, K8sConfigCreate,
    PrometheusConfig, PrometheusConfigCreate,
    NamespaceQuota, SyncResult
)
from app.storage import Storage, ReportStorage, MemoryReportCache, K8sConfigStorage, PrometheusConfigStorage, NamespaceQuotaStorage, SchedulerConfigStorage
from app.parser import ReportParser
from app.quota_manager import QuotaManager
from app.calculator import Calculator
from app.exporter import ExcelExporter


BASE_DIR = Path(__file__).parent.parent
storage = ReportStorage(str(BASE_DIR / "uploads"))
quota_storage = Storage(str(BASE_DIR / "data"))
report_cache = MemoryReportCache()
k8s_config_storage = K8sConfigStorage(str(BASE_DIR / "data"))
prometheus_config_storage = PrometheusConfigStorage(str(BASE_DIR / "data"))
namespace_quota_storage = NamespaceQuotaStorage(str(BASE_DIR / "data"))
scheduler_config_storage = SchedulerConfigStorage(str(BASE_DIR / "data"))

quota_manager = QuotaManager(quota_storage)
calculator = Calculator(quota_storage)
exporter = ExcelExporter(calculator)

app = FastAPI(title="K8s Resource Analyzer", debug=True)

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)

app.mount("/static", StaticFiles(directory=str(BASE_DIR / "static")), name="static")
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))


@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})


@app.get("/api/templates/quota")
async def download_quota_template():
    from openpyxl import Workbook
    
    wb = Workbook()
    ws = wb.active
    ws.title = "配额模板"
    
    headers = ["云序号", "项目名称", "CPU配额(核)", "内存配额(GB)"]
    ws.append(headers)
    
    ws.append(["inst_example1", "示例项目A", 10, 50])
    ws.append(["inst_example2", "示例项目B", 20, 100])
    
    for col in ['A', 'B', 'C', 'D']:
        ws.column_dimensions[col].width = 20
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=quota_template.xlsx"}
    )


@app.get("/api/templates/report")
async def download_report_template():
    content = """项目名称;命名空间名称;CPU使用总量(核);内存使用总量(bytes)
示例项目A;namespace-a;3.5;16106127360
示例项目A;namespace-b;2.1;9663676416
示例项目B;namespace-c;5.0;21474836480"""
    
    return Response(
        content=content.encode("utf-8"),
        media_type="text/plain",
        headers={"Content-Disposition": "attachment; filename=Day_report_TEMPLATE.txt"}
    )


@app.get("/api/quotas", response_model=List[Quota])
async def get_quotas():
    return quota_manager.get_all_quotas()


@app.post("/api/quotas")
async def create_quota(quota: Quota):
    success = quota_manager.add_quota(
        quota.cloud_id, quota.project_name, 
        quota.cpu_quota, quota.memory_quota
    )
    if not success:
        raise HTTPException(status_code=400, detail="Cloud ID already exists")
    return {"success": True}


@app.put("/api/quotas/{cloud_id}")
async def update_quota(cloud_id: str, update: QuotaUpdate):
    result = quota_manager.update_quota(cloud_id, update)
    if not result:
        raise HTTPException(status_code=404, detail="Quota not found")
    return result


@app.delete("/api/quotas/{cloud_id}")
async def delete_quota(cloud_id: str):
    success = quota_manager.delete_quota(cloud_id)
    if not success:
        raise HTTPException(status_code=404, detail="Quota not found")
    return {"success": True}


@app.post("/api/quotas/import", response_model=ImportResult)
async def import_quotas(file: UploadFile = File(...)):
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        content = await file.read()
        tmp.write(content)
        tmp_path = tmp.name

    try:
        result = quota_manager.import_from_excel(tmp_path)
        return result
    finally:
        os.unlink(tmp_path)


@app.post("/api/reports/upload")
async def upload_reports(files: List[UploadFile] = File(...)):
    uploaded = []
    errors = []

    for file in files:
        try:
            date = ReportParser.extract_date_from_filename(file.filename)
            if not date:
                errors.append(f"{file.filename}: 文件名中未找到日期，格式应为 YYYY-MM-DD")
                continue

            content = await file.read()
            text_content = content.decode("utf-8")

            project_data = ReportParser.parse_file(text_content)
            project_usages = calculator.process_report(project_data)

            report_cache.add(date, project_usages)
            uploaded.append({
                "filename": file.filename,
                "date": date,
                "record_count": len(project_usages)
            })

        except Exception as e:
            errors.append(f"{file.filename}: {str(e)}")

    return {"uploaded": uploaded, "errors": errors}


@app.get("/api/reports")
async def get_reports():
    dates = report_cache.get_dates()
    result = []
    for d in dates:
        usages = report_cache.get(d) or []
        with_cloud = len([u for u in usages if u.cloud_id is not None])
        result.append({"date": d, "record_count": len(usages), "with_cloud_id": with_cloud})
    return result


@app.get("/api/reports/export")
async def export_report(dates: str = Query(...)):
    date_list = [d.strip() for d in dates.split(",")]
    reports = []
    
    all_dates = report_cache.get_dates()
    
    for date in date_list:
        usages = report_cache.get(date)
        if usages:
            filtered = [u for u in usages if u.cloud_id is not None]
            if filtered:
                reports.append((date, filtered))
    
    if not reports:
        raise HTTPException(
            status_code=404, 
            detail=f"未找到有效报告。请求的日期: {date_list}, 可用日期: {all_dates}"
        )

    excel_data = exporter.export_multiple(reports)

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(excel_data)
        tmp_path = tmp.name

    async def cleanup():
        import os
        try:
            os.unlink(tmp_path)
        except:
            pass

    return FileResponse(
        path=tmp_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="resource_usage_report.xlsx"
    )


@app.get("/api/reports/{date}", response_model=List[ProjectUsage])
async def get_report(date: str):
    usages = report_cache.get(date)
    if usages is None:
        raise HTTPException(status_code=404, detail="Report not found")
    filtered = [u for u in usages if u.cloud_id is not None]
    return filtered


@app.get("/api/k8s/configs", response_model=List[K8sConfig])
async def get_k8s_configs():
    return k8s_config_storage.get_all()


@app.post("/api/k8s/configs", response_model=K8sConfig)
async def create_k8s_config(config: K8sConfigCreate):
    import uuid
    new_config = K8sConfig(
        id=str(uuid.uuid4()),
        name=config.name,
        kubeconfig=config.kubeconfig,
        use_projectquota=config.use_projectquota
    )
    if not k8s_config_storage.add(new_config):
        raise HTTPException(status_code=400, detail="Config ID already exists")
    return new_config


@app.delete("/api/k8s/configs/{config_id}")
async def delete_k8s_config(config_id: str):
    if not k8s_config_storage.delete(config_id):
        raise HTTPException(status_code=404, detail="Config not found")
    return {"success": True}


@app.post("/api/k8s/sync", response_model=SyncResult)
async def sync_k8s_config(config_id: str):
    config = k8s_config_storage.get(config_id)
    if not config:
        raise HTTPException(status_code=404, detail="Config not found")
    
    from app.k8s_client import sync_namespace_quota_from_k8s
    result = sync_namespace_quota_from_k8s(config.kubeconfig, config.name, namespace_quota_storage)
    return result


@app.post("/api/k8s/sync-project-quota", response_model=SyncResult)
async def sync_project_quota(config_id: str):
    config = k8s_config_storage.get(config_id)
    if not config:
        raise HTTPException(status_code=404, detail="Config not found")
    
    from app.k8s_client import sync_k8s_quotas
    result = sync_k8s_quotas(config.kubeconfig, quota_storage, use_projectquota=True)
    return result


@app.get("/api/k8s/quotas", response_model=List[NamespaceQuota])
async def get_namespace_quotas(cluster: str = None):
    if cluster:
        return namespace_quota_storage.get_by_cluster(cluster)
    return namespace_quota_storage.get_all()


@app.delete("/api/k8s/quotas/{cluster_name}/{namespace}")
async def delete_namespace_quota(cluster_name: str, namespace: str):
    if not namespace_quota_storage.delete(cluster_name, namespace):
        raise HTTPException(status_code=404, detail="Namespace quota not found")
    return {"success": True}


@app.post("/api/k8s/quotas/import", response_model=ImportResult)
async def import_namespace_quotas(file: UploadFile = File(...)):
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        content = await file.read()
        tmp.write(content)
        tmp_path = tmp.name

    try:
        from openpyxl import load_workbook
        wb = load_workbook(tmp_path)
        ws = wb.active
        
        imported = 0
        errors = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                cluster_name, namespace_name, project_name, cpu_limit, memory_limit, cpu_used, memory_used = row[:7]
                if not cluster_name or not namespace_name:
                    errors.append(f"第{row_idx}行: 缺少集群名称或命名空间")
                    continue
                
                ns_quota = NamespaceQuota(
                    cluster_name=str(cluster_name),
                    namespace=str(namespace_name),
                    project_name=str(project_name) if project_name else "",
                    cpu_limit=float(cpu_limit) if cpu_limit else 0,
                    memory_limit=float(memory_limit) if memory_limit else 0,
                    cpu_used=float(cpu_used) if cpu_used else 0,
                    memory_used=float(memory_used) if memory_used else 0
                )
                namespace_quota_storage.save(ns_quota)
                imported += 1
            except Exception as e:
                errors.append(f"第{row_idx}行: {str(e)}")
        
        return ImportResult(success=True, imported=imported, updated=0, errors=errors)
    except Exception as e:
        return ImportResult(success=False, imported=0, updated=0, errors=[f"读取文件失败: {str(e)}"])
    finally:
        os.unlink(tmp_path)


@app.get("/api/templates/namespace-quota")
async def download_namespace_quota_template():
    from openpyxl import Workbook
    
    wb = Workbook()
    ws = wb.active
    ws.title = "命名空间配额模板"
    
    headers = ["集群名称", "命名空间", "项目名称", "CPU配额(核)", "内存配额(GB)", "CPU已用(核)", "内存已用(GB)"]
    ws.append(headers)
    
    ws.append(["prod-cluster", "namespace-a", "项目A", 100, 200, 50, 100])
    ws.append(["prod-cluster", "namespace-b", "项目B", 50, 100, 20, 40])
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 18
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=namespace_quota_template.xlsx"}
    )


@app.get("/api/prometheus/configs", response_model=List[PrometheusConfig])
async def get_prometheus_configs():
    return prometheus_config_storage.get_all()


@app.post("/api/prometheus/configs", response_model=PrometheusConfig)
async def create_prometheus_config(config: PrometheusConfigCreate):
    import uuid
    new_config = PrometheusConfig(
        id=str(uuid.uuid4()),
        name=config.name,
        url=config.url,
        cluster_name=config.cluster_name,
        use_accurate_sync=config.use_accurate_sync
    )
    if not prometheus_config_storage.add(new_config):
        raise HTTPException(status_code=400, detail="Config ID already exists")
    return new_config


@app.delete("/api/prometheus/configs/{config_id}")
async def delete_prometheus_config(config_id: str):
    if not prometheus_config_storage.delete(config_id):
        raise HTTPException(status_code=404, detail="Config not found")
    return {"success": True}


@app.post("/api/prometheus/sync", response_model=SyncResult)
async def sync_prometheus_config(config_id: str, date: str = None):
    from datetime import datetime, timedelta
    
    config = prometheus_config_storage.get(config_id)
    if not config:
        raise HTTPException(status_code=404, detail="Config not found")
    
    from app.prometheus_client import sync_prometheus_usage_accurate
    k8s_config = k8s_config_storage.get_all()
    matching_k8s = next((k for k in k8s_config if k.name == config.cluster_name), None)
    
    if not matching_k8s:
        raise HTTPException(status_code=400, detail="No matching K8s config found for this cluster")
    
    if not date:
        date = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
    
    result = sync_prometheus_usage_accurate(
        config.url,
        matching_k8s.kubeconfig,
        config.cluster_name,
        date,
        namespace_quota_storage,
        report_cache
    )
    return result


@app.post("/api/prometheus/import", response_model=ImportResult)
async def import_prometheus_configs(file: UploadFile = File(...)):
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        content = await file.read()
        tmp.write(content)
        tmp_path = tmp.name

    try:
        from openpyxl import load_workbook
        wb = load_workbook(tmp_path)
        ws = wb.active
        
        imported = 0
        errors = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                name, url, cluster_name, use_accurate_sync = row
                if not name or not url or not cluster_name:
                    errors.append(f"第{row_idx}行: 缺少必填字段")
                    continue
                
                import uuid
                new_config = PrometheusConfig(
                    id=str(uuid.uuid4()),
                    name=str(name),
                    url=str(url),
                    cluster_name=str(cluster_name),
                    use_accurate_sync=bool(use_accurate_sync) if use_accurate_sync is not None else False
                )
                
                if prometheus_config_storage.add(new_config):
                    imported += 1
                else:
                    errors.append(f"第{row_idx}行: 配置已存在")
            except Exception as e:
                errors.append(f"第{row_idx}行: {str(e)}")
        
        return ImportResult(success=True, imported=imported, updated=0, errors=errors)
    except Exception as e:
        return ImportResult(success=False, imported=0, updated=0, errors=[f"读取文件失败: {str(e)}"])
    finally:
        os.unlink(tmp_path)


@app.get("/api/templates/prometheus-config")
async def download_prometheus_config_template():
    from openpyxl import Workbook
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Prometheus配置模板"
    
    headers = ["配置名称", "Prometheus URL", "集群名称", "精确同步"]
    ws.append(headers)
    
    ws.append(["prod-prometheus", "http://prometheus:9090", "prod-cluster", True])
    ws.append(["dev-prometheus", "http://prometheus-dev:9090", "dev-cluster", False])
    
    for col in ['A', 'B', 'C', 'D']:
        ws.column_dimensions[col].width = 25
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=prometheus_config_template.xlsx"}
    )


@app.get("/api/scheduler/config")
async def get_scheduler_config():
    return scheduler_config_storage.get()


@app.post("/api/scheduler/config")
async def update_scheduler_config(
    enabled: bool,
    namespace_quota_hour: int = 2,
    prometheus_sync_hour: int = 3,
    interval_hours: int = 24
):
    return scheduler_config_storage.update(enabled, namespace_quota_hour, prometheus_sync_hour, interval_hours)


@app.get("/api/templates/k8s-config")
async def download_k8s_config_template():
    from openpyxl import Workbook
    
    wb = Workbook()
    ws = wb.active
    ws.title = "K8s集群配置模板"
    
    headers = ["集群名称", "Kubeconfig (Base64编码或原始内容)", "使用ProjectQuota"]
    ws.append(headers)
    
    ws.append(["prod-cluster", "YXdzOi4uLg==", True])
    ws.append(["dev-cluster", "YXdzOi4uLg==", False])
    
    for col in ['A', 'B', 'C']:
        ws.column_dimensions[col].width = 30
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=k8s_config_template.xlsx"}
    )


@app.post("/api/k8s/configs/import", response_model=ImportResult)
async def import_k8s_configs(file: UploadFile = File(...)):
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        content = await file.read()
        tmp.write(content)
        tmp_path = tmp.name

    try:
        from openpyxl import load_workbook
        wb = load_workbook(tmp_path)
        ws = wb.active
        
        imported = 0
        errors = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                name, kubeconfig, use_projectquota = row[:3]
                if not name or not kubeconfig:
                    errors.append(f"第{row_idx}行: 缺少集群名称或kubeconfig")
                    continue
                
                import uuid
                new_config = K8sConfig(
                    id=str(uuid.uuid4()),
                    name=str(name),
                    kubeconfig=str(kubeconfig),
                    use_projectquota=bool(use_projectquota) if use_projectquota is not None else True
                )
                
                if k8s_config_storage.add(new_config):
                    imported += 1
                else:
                    errors.append(f"第{row_idx}行: 配置已存在")
            except Exception as e:
                errors.append(f"第{row_idx}行: {str(e)}")
        
        return ImportResult(success=True, imported=imported, updated=0, errors=errors)
    except Exception as e:
        return ImportResult(success=False, imported=0, updated=0, errors=[f"读取文件失败: {str(e)}"])
    finally:
        os.unlink(tmp_path)
