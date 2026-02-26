from typing import List, Optional
from openpyxl import load_workbook

from app.models import Quota, QuotaUpdate, ImportResult
from app.storage import Storage


class QuotaManager:
    def __init__(self, storage: Storage):
        self.storage = storage

    def get_all_quotas(self) -> List[Quota]:
        return self.storage.get_all_quotas()

    def get_quota(self, cloud_id: str) -> Optional[Quota]:
        return self.storage.get_quota(cloud_id)

    def add_quota(self, cloud_id: str, project_name: str, cpu_quota: float, memory_quota: float) -> bool:
        quota = Quota(
            cloud_id=cloud_id,
            project_name=project_name,
            cpu_quota=cpu_quota,
            memory_quota=memory_quota
        )
        return self.storage.add_quota(quota)

    def update_quota(self, cloud_id: str, update: QuotaUpdate) -> Optional[Quota]:
        existing = self.storage.get_quota(cloud_id)
        if not existing:
            return None

        quota = Quota(
            cloud_id=cloud_id,
            project_name=update.project_name if update.project_name is not None else existing.project_name,
            cpu_quota=update.cpu_quota if update.cpu_quota is not None else existing.cpu_quota,
            memory_quota=update.memory_quota if update.memory_quota is not None else existing.memory_quota
        )
        self.storage.update_quota(cloud_id, quota)
        return quota

    def delete_quota(self, cloud_id: str) -> bool:
        return self.storage.delete_quota(cloud_id)

    def import_from_excel(self, filepath: str) -> ImportResult:
        imported = 0
        updated = 0
        errors = []

        try:
            wb = load_workbook(filepath, read_only=True)
            ws = wb.active

            rows = list(ws.iter_rows(min_row=2, values_only=True))

            quotas = []
            for idx, row in enumerate(rows, start=2):
                if not row[0] and not row[1]:
                    continue

                try:
                    cloud_id = str(row[0]).strip() if row[0] else ""
                    project_name = str(row[1]).strip() if row[1] else ""
                    cpu_quota = float(row[2]) if row[2] is not None else 0.0
                    memory_quota = float(row[3]) if row[3] is not None else 0.0

                    if not cloud_id:
                        errors.append(f"Row {idx}: 云序号不能为空")
                        continue

                    if not project_name:
                        errors.append(f"Row {idx}: 项目名称不能为空")
                        continue

                    quotas.append(Quota(
                        cloud_id=cloud_id,
                        project_name=project_name,
                        cpu_quota=cpu_quota,
                        memory_quota=memory_quota
                    ))

                except (ValueError, TypeError) as e:
                    errors.append(f"Row {idx}: 数据格式错误 - {str(e)}")

            imported, updated = self.storage.import_quotas(quotas)

            wb.close()

        except Exception as e:
            errors.append(f"Excel读取失败: {str(e)}")

        return ImportResult(
            success=len(errors) == 0,
            imported=imported,
            updated=updated,
            errors=errors
        )
